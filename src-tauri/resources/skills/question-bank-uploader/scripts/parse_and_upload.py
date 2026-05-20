#!/usr/bin/env python3
"""
题库解析和上传脚本
用法: python3 parse_and_upload.py <xlsx_file> --api-url <url> [--token <token>]
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("错误: 需要安装 requests 库")
    print("请运行: pip install requests")
    sys.exit(1)


def parse_xlsx(file_path: str) -> list[dict]:
    """解析 xlsx 文件，返回题目列表"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    questions = []
    errors = []

    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 验证必需列
    required_cols = ["题目类型", "题目内容", "正确答案"]
    for col in required_cols:
        if col not in headers:
            errors.append(f"缺少必需列: {col}")

    if errors:
        return {"errors": errors, "questions": []}

    # 获取列索引
    col_map = {h: i for i, h in enumerate(headers)}

    # 解析数据行
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = [cell.value for cell in row]

        # 跳过空行
        if not any(row_data):
            continue

        try:
            question = {
                "type": row_data[col_map.get("题目类型", 0)] or "choice",
                "content": row_data[col_map.get("题目内容", 1)] or "",
                "options": {
                    "A": row_data[col_map.get("选项A", 2)] or "",
                    "B": row_data[col_map.get("选项B", 3)] or "",
                    "C": row_data[col_map.get("选项C", 4)] or "",
                    "D": row_data[col_map.get("选项D", 5)] or "",
                },
                "answer": row_data[col_map.get("正确答案", 6)] or "",
                "explanation": row_data[col_map.get("解析", 7)] or "",
            }

            # 验证必填字段
            if not question["content"]:
                errors.append(f"第 {row_num} 行: 题目内容为空")
                continue

            if not question["answer"]:
                errors.append(f"第 {row_num} 行: 正确答案为空")
                continue

            questions.append(question)

        except Exception as e:
            errors.append(f"第 {row_num} 行: 解析错误 - {str(e)}")

    return {"questions": questions, "errors": errors}


def upload_questions(questions: list[dict], api_url: str, token: str = None) -> dict:
    """上传题目到 API"""
    url = f"{api_url}/api/questions/batch"

    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    try:
        response = requests.post(url, json=questions, headers=headers, timeout=30)
        response.raise_for_status()
        return {"success": True, "data": response.json()}
    except requests.exceptions.RequestException as e:
        return {"success": False, "error": str(e)}


def main():
    parser = argparse.ArgumentParser(description="题库解析和上传工具")
    parser.add_argument("file", help="xlsx 文件路径")
    parser.add_argument("--api-url", required=True, help="API 服务器地址")
    parser.add_argument("--token", help="API 认证 token")
    parser.add_argument("--dry-run", action="store_true", help="仅解析，不上传")

    args = parser.parse_args()

    # 检查文件是否存在
    if not Path(args.file).exists():
        print(f"错误: 文件不存在 - {args.file}")
        sys.exit(1)

    # 解析 xlsx
    print(f"正在解析: {args.file}")
    result = parse_xlsx(args.file)

    if result["errors"]:
        print("\n⚠️  解析警告:")
        for error in result["errors"]:
            print(f"  - {error}")

    if not result["questions"]:
        print("\n❌ 没有找到有效的题目")
        sys.exit(1)

    print(f"\n✅ 解析完成: 共 {len(result['questions'])} 道题目")

    # 显示前几道题目预览
    print("\n📋 题目预览:")
    for i, q in enumerate(result["questions"][:3], 1):
        print(f"  {i}. [{q['type']}] {q['content'][:50]}...")
    if len(result["questions"]) > 3:
        print(f"  ... 还有 {len(result['questions']) - 3} 道题目")

    # 上传
    if args.dry_run:
        print("\n🔍 Dry run 模式，跳过上传")
        print(json.dumps(result["questions"], ensure_ascii=False, indent=2))
        return

    print(f"\n📤 正在上传到: {args.api_url}")
    upload_result = upload_questions(result["questions"], args.api_url, args.token)

    if upload_result["success"]:
        print("✅ 上传成功!")
        print(json.dumps(upload_result["data"], ensure_ascii=False, indent=2))
    else:
        print(f"❌ 上传失败: {upload_result['error']}")
        sys.exit(1)


if __name__ == "__main__":
    main()
